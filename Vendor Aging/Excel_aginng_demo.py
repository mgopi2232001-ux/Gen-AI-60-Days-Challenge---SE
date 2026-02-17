import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox


# ================== CORE AGING LOGIC ================== #

def create_vendor_aging(
    file_path: str,
    sheet_name: str | None = None,
    aging_sheet: str = "aging",
    as_of_date: str | None = None,
):
    """
    Vendor-wise aging based on INVOICE DATE & AMOUNT.

    file_path   : path to the Excel file
    sheet_name  : sheet to read from; if None -> first sheet
    aging_sheet : sheet to create/replace with vendor-wise aging
    as_of_date  : 'YYYY-MM-DD'. If None -> today's date.

    Expected columns in the source sheet:
        - Vendor Name
        - Invoice Date
        - Amount
    """

    # If no as-of date given, use today
    if not as_of_date:
        as_of_date = datetime.today().strftime("%Y-%m-%d")

    # ---------- Step 1: Pick sheet ----------
    xl = pd.ExcelFile(file_path)
    if sheet_name is None:
        sheet_name = xl.sheet_names[0]   # just take first sheet

    # ---------- Step 2: Read data ----------
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        raise ValueError(f"Error reading sheet '{sheet_name}' from file.\n{e}")

    # Check required columns
    required_cols = ["Vendor Name", "Invoice Date", "Amount"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(
            f"Missing required column(s): {missing}\n"
            f"Your sheet must contain at least: {required_cols}"
        )

    # ---------- Step 3: Clean & compute Age ----------
    # Parse dates (dayfirst=True to handle formats like 12-11-2025)
    df["Invoice Date"] = pd.to_datetime(
        df["Invoice Date"],
        dayfirst=True,
        errors="coerce"
    )

    # Drop rows where date or amount is invalid
    df = df.dropna(subset=["Invoice Date", "Amount"]).copy()

    # Convert as-of date
    as_of = pd.to_datetime(as_of_date)
    df["Age (Days)"] = (as_of - df["Invoice Date"]).dt.days
    df["Age (Days)"] = df["Age (Days)"].clip(lower=0)  # future dates -> 0

    # ---------- Step 4: Define buckets ----------
    # <30   : 0–29
    # <60   : 30–59
    # <90   : 60–89
    # <180  : 90–179
    # >180  : 180+
    bins = [-1, 29, 59, 89, 179, 10**9]
    labels = ["<30", "<60", "<90", "<180", ">180"]

    df["Bucket"] = pd.cut(
        df["Age (Days)"],
        bins=bins,
        labels=labels,
        right=True,
        include_lowest=True,
    )

    bucket_order = ["<30", "<60", "<90", "<180", ">180"]
    df["Bucket"] = pd.Categorical(df["Bucket"], categories=bucket_order, ordered=True)

    # ---------- Step 5: Vendor-wise summary (ONLY Vendor Name) ----------
    summary = (
        df
        .groupby(["Vendor Name", "Bucket"], as_index=False)["Amount"]
        .sum()
        .pivot_table(
            index=["Vendor Name"],
            columns="Bucket",
            values="Amount",
            fill_value=0,
        )
    )

    # Ensure fixed column order
    summary = summary.reindex(columns=bucket_order, fill_value=0)

    # Add total column
    summary["Total"] = summary.sum(axis=1)
    summary = summary.reset_index()

    # ---------- Step 6: Write to Excel (aging sheet) ----------
    # Remove existing aging sheet (if present)
    book = load_workbook(file_path)
    if aging_sheet in book.sheetnames:
        ws = book[aging_sheet]
        book.remove(ws)
        book.save(file_path)

    # Append new aging sheet
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
        summary.to_excel(writer, sheet_name=aging_sheet, index=False, startrow=0)

    return sheet_name, as_of_date


# ================== TKINTER GUI ================== #

def browse_file():
    path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel Files", "*.xlsx *.xls")],
    )
    if path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, path)


def run_aging():
    file_path = file_entry.get().strip()
    as_of_date = date_entry.get().strip()

    if not file_path:
        messagebox.showerror("Error", "Please select or enter an Excel file path.")
        return

    try:
        # Validate date format if user typed something
        if as_of_date:
            datetime.strptime(as_of_date, "%Y-%m-%d")

        used_sheet, used_date = create_vendor_aging(
            file_path=file_path,
            sheet_name=None,          # use first sheet
            as_of_date=as_of_date or None,
        )

        messagebox.showinfo(
            "Success",
            f"Vendor-wise aging created successfully!\n\n"
            f"Source sheet : {used_sheet}\n"
            f"As-of Date   : {used_date}\n\n"
            f"Open the same workbook and check the 'aging' sheet."
        )

    except PermissionError:
        messagebox.showerror(
            "Permission Error",
            "Windows blocked access to this file.\n\n"
            "Most common reasons:\n"
            "• The Excel file is currently OPEN.\n"
            "• OneDrive/SharePoint is still syncing.\n\n"
            "Close the file in Excel, wait a few seconds, and try again.",
        )
    except ValueError as ve:
        messagebox.showerror("Error", str(ve))
    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n{e}")


# ================== BUILD WINDOW ================== #

root = tk.Tk()
root.title("Vendor Aging (Vendor-wise)")

root.geometry("540x190")

# File path row
tk.Label(root, text="Excel File Path:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
file_entry = tk.Entry(root, width=50)
file_entry.grid(row=0, column=1, padx=5, pady=10, sticky="w")
browse_btn = tk.Button(root, text="Browse", command=browse_file)
browse_btn.grid(row=0, column=2, padx=5, pady=10)

# As-of Date row
tk.Label(root, text="As-of Date (YYYY-MM-DD):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
date_entry = tk.Entry(root, width=20)
date_entry.grid(row=1, column=1, padx=5, pady=10, sticky="w")
date_entry.insert(0, datetime.today().strftime("%Y-%m-%d"))

# Run button
run_btn = tk.Button(root, text="Create Vendor Aging", command=run_aging, width=22)
run_btn.grid(row=2, column=1, padx=5, pady=20)

root.mainloop()
# ================== END OF SCRIPT ================== #